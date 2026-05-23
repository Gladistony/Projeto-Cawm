"""
Seed específico da bacia Poço do Fumo (BRÍGIDA).

Este script:
- lê os parâmetros iniciais do XLSM
- lê evaporação mensal do TXT
- lê precipitação diária do TXT
- lê vazão diária do TXT
- insere os períodos de calibração/validação correspondentes
- insere os resultados de desempenho correspondentes
- grava tudo no banco com upsert

Ainda não executa calibração; apenas faz a carga dos dados dessa bacia.
"""
from __future__ import annotations

import re
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import cast

import pandas as pd
from openpyxl import load_workbook
from db_init import initialize_db
from db_models import (
    Bacia,
    CalibrationPeriod,
    ModelResult,
    EvaporationMonthly,
    PrecipitationDaily,
    FlowDaily,
)

NOME_BACIA = "Poço do Fumo"
REGIAO = "BRIGIDA"
BASE_DIR = Path("Dados/Brigida")
XLSM_PATH = BASE_DIR / "CAWM_2023_POCO_DO_FUMO_TA.xlsm"
EVAP_PATH = BASE_DIR / "Evap_Poco_do_Fumo.txt"
PLU_PATH = BASE_DIR / "Plu_Poco_do_Fumo.txt"
FLU_PATH = BASE_DIR / "Flu_Poco_do_Fumo.txt"
PERIODOS_CSV = Path("Tabela 2: Períodos de calibração e validação.csv")
RESULTADOS_CSV = Path("Tabela 3: Parâmetros do modelo e coeficientes de desempenho_cleaned.csv")


PARAMETER_MAP = {
    "Área da bacia - Ab (km²)": "area_km2",
    "Parâmetro do escoamento na calha (calculado) - K": "k",
    "Coeficiente da função de evapotranspiração - α (padrão 1,4)": "a",
    "Coeficiente do expoente das perdas na calha (entre 1 e 2) - p": "expo_perdas",
    "Coeficiente do expoente do escoamento na calha (≈5/3) - b": "b",
    "Capacidade de armazenamento percolação profunda - Gmax ": "gmax",
    "Capacidade de armazenamento no solo - S - SUBmax": "submax",
    "Reserva no solo inicial": "reserva_solo_inicial",
    "Profundo corrigido inicial": "profundo_inicial",
    "Reserva na calha do rio inicial": "s3_inicial",
    "Rios temporários (0) ou perenes (1)": "rio",
}


RESULTADOS_COLUMNS = {
    "Area (km²)": "area_km2",
    "S (mm)": "s_mm",
    "Ks": "ks",
    "NSE Cal.": "nse_calib",
    "NSE Val.": "nse_val",
    "NSEsqrt Cal.": "nse_sqrt_calib",
    "NSEsqrt Val.": "nse_sqrt_val",
    "NSElog Cal.": "nse_log_calib",
    "NSElog Val.": "nse_log_val",
    "Pbias (%) Cal.": "pbias_calib",
    "Pbias (%) Val.": "pbias_val",
}

MONTH_MAP = {
    "jan": 1,
    "fev": 2,
    "mar": 3,
    "abr": 4,
    "mai": 5,
    "jun": 6,
    "jul": 7,
    "ago": 8,
    "set": 9,
    "out": 10,
    "nov": 11,
    "dez": 12,
}


DATE_RE = re.compile(r"(\d{1,2}/\d{1,2}/\d{2,4})")


def normalize_text(value: str) -> str:
    if not isinstance(value, str):
        return str(value).upper().strip()
    normalized = ''.join(
        character for character in unicodedata.normalize("NFD", value)
        if unicodedata.category(character) != "Mn"
    )
    return normalized.upper().strip()


def to_float(value):
    if value is None or value == "":
        return None
    try:
        return float(value)
    except Exception:
        return None


def to_int(value):
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except Exception:
        return None


def parse_date(value: str):
    for fmt in ("%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(value, fmt).date()
        except Exception:
            continue
    return None


def parse_period(value: str):
    found = DATE_RE.findall(value or "")
    if len(found) < 2:
        return None, None
    return parse_date(found[0]), parse_date(found[1])


def find_numeric_in_row(ws, row_number: int, start_col: int = 2):
    for col in range(start_col, ws.max_column + 1):
        value = ws.cell(row_number, col).value
        if isinstance(value, (int, float)):
            return value
    return None


def find_row_by_label(ws, label_text: str):
    normalized_target = normalize_text(label_text)
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and normalize_text(cell.value) == normalized_target:
                return cell.row
    return None


def extract_bacia_parameters(xlsm_path: Path) -> tuple[dict[str, object], list[str]]:
    workbook = load_workbook(xlsm_path, data_only=True, read_only=True)
    sheet = workbook["SetUp"]

    parametros: dict[str, object] = {
        "nome": NOME_BACIA,
        "regiao": REGIAO,
    }

    missing = []
    for label, field in PARAMETER_MAP.items():
        row_number = find_row_by_label(sheet, label)
        if row_number is None:
            missing.append(field)
            continue

        value = find_numeric_in_row(sheet, row_number)
        if value is None:
            missing.append(field)
            continue

        if field in {"rio"}:
            parametros[field] = to_int(value)
        elif field == "area_km2":
            parametros[field] = to_float(value)
        else:
            parametros[field] = to_float(value)

    # Padrões do sistema em funcoes.py
    if parametros.get("kg") is None:
        parametros["kg"] = 1.0
        if "kg" in missing:
            missing.remove("kg")

    if parametros.get("beta") is None:
        parametros["beta"] = 2.0
        if "beta" in missing:
            missing.remove("beta")

    workbook.close()
    return parametros, missing


def read_monthly_evaporation(path: Path) -> list[dict]:
    frame = pd.read_csv(path, sep=r"\s+", engine="python", header=None)
    if frame.shape[1] < 2:
        raise ValueError(f"Evaporação inválida em {path}")

    clean = []
    for raw_month, raw_value in frame.iloc[:, :2].itertuples(index=False):
        if pd.isna(raw_month) or pd.isna(raw_value):
            continue
        month_text = normalize_text(raw_month).lower()[:3]
        if month_text not in MONTH_MAP:
            continue
        clean.append({"mes": MONTH_MAP[month_text], "valor": round(float(raw_value), 3)})
    return clean


def read_daily_series(path: Path) -> list[dict]:
    frame = pd.read_csv(path, sep=r"\s+", engine="python", header=None)
    if frame.shape[1] < 2:
        raise ValueError(f"Série diária inválida em {path}")

    series = []
    for raw_date, raw_value in frame.iloc[:, :2].itertuples(index=False):
        if pd.isna(raw_date) or pd.isna(raw_value):
            continue
        parsed_date = parse_date(str(raw_date))
        if parsed_date is None:
            continue
        series.append({"data": parsed_date, "valor": round(float(raw_value), 3)})
    return series


def load_periods(csv_path: Path):
    df = pd.read_csv(csv_path, sep=';', dtype=str, keep_default_na=False)
    df = df[(df["Bacia"].str.upper() == "BRÍGIDA") & (df["Estação Fluviométrica"].str.contains("Poço do Fumo", case=False, na=False))]
    return df


def load_results(csv_path: Path):
    df = pd.read_csv(csv_path, dtype=str, keep_default_na=False)
    df = df[(df["Bacia"].str.upper() == "BRÍGIDA") & (df["Flu. Station"].str.contains("Poço do Fumo", case=False, na=False))]
    return df


def upsert_bacia(session, parametros: dict) -> Bacia:
    existing = None
    target_name = normalize_text(NOME_BACIA)
    for bacia in session.query(Bacia).all():
        if normalize_text(bacia.nome) == target_name:
            existing = bacia
            break
    if existing:
        for key, value in parametros.items():
            if hasattr(existing, key) and value is not None:
                setattr(existing, key, value)
        return existing

    bacia = Bacia(**parametros)
    session.add(bacia)
    session.flush()
    return bacia


def upsert_monthly(session, bacia_id: int, items: list[dict]) -> tuple[int, int]:
    created = 0
    updated = 0
    for item in items:
        existing = session.query(EvaporationMonthly).filter_by(bacia_id=bacia_id, mes=item["mes"]).first()
        if existing:
            existing.valor = item["valor"]
            updated += 1
        else:
            session.add(EvaporationMonthly(bacia_id=bacia_id, mes=item["mes"], valor=item["valor"]))
            created += 1
    return created, updated


def upsert_daily(session, model, bacia_id: int, items: list[dict]) -> tuple[int, int]:
    created = 0
    updated = 0
    for item in items:
        existing = session.query(model).filter_by(bacia_id=bacia_id, data=item["data"]).first()
        if existing:
            existing.valor = item["valor"]
            updated += 1
        else:
            session.add(model(bacia_id=bacia_id, data=item["data"], valor=item["valor"]))
            created += 1
    return created, updated


def upsert_periods(session, bacia_id: int, periods_df: pd.DataFrame):
    created = 0
    updated = 0
    for _, row in periods_df.iterrows():
        station = row["Estação Fluviométrica"].strip()
        method = "normal"
        if "(" in station and ")" in station:
            suffix = station[station.find("(") + 1: station.find(")")].lower().replace(" ", "")
            if "c" in suffix and "r" in suffix:
                method = "c.r"
            elif "c" in suffix:
                method = "c"
            elif "r" in suffix:
                method = "r"
            station = station[:station.find("(")].strip()

        calib_start, calib_end = parse_period(row["Período de Calibração"])
        val_start, val_end = parse_period(row["Período de Validação"])

        existing = session.query(CalibrationPeriod).filter_by(bacia_id=bacia_id, station=station, method=method).first()
        if existing:
            existing.calib_start = calib_start
            existing.calib_end = calib_end
            existing.val_start = val_start
            existing.val_end = val_end
            updated += 1
        else:
            session.add(
                CalibrationPeriod(
                    bacia_id=bacia_id,
                    station=station,
                    method=method,
                    calib_start=calib_start,
                    calib_end=calib_end,
                    val_start=val_start,
                    val_end=val_end,
                )
            )
            created += 1
    return created, updated


def upsert_results(session, bacia_id: int, resultados_df: pd.DataFrame):
    created = 0
    updated = 0
    periods = session.query(CalibrationPeriod).filter_by(bacia_id=bacia_id).all()

    def find_period_id(station_name: str):
        clean = station_name.strip()
        method = "normal"
        if "(" in clean and ")" in clean:
            suffix = clean[clean.find("(") + 1: clean.find(")")].lower().replace(" ", "")
            if "c" in suffix and "r" in suffix:
                method = "c.r"
            elif "c" in suffix:
                method = "c"
            elif "r" in suffix:
                method = "r"
            clean = clean[:clean.find("(")].strip()

        for period in periods:
            period_station = period.station or ""
            if normalize_text(period_station) == normalize_text(clean) and period.method == method:
                return period.id
            if normalize_text(period_station) == normalize_text(clean):
                return period.id
        return None

    for _, row in resultados_df.iterrows():
        station = row["Flu. Station"].strip()
        period_id = find_period_id(station)
        if period_id is None:
            continue

        payload = {
            key: (float(row[col]) if row[col] not in ("", None) else None)
            for col, key in RESULTADOS_COLUMNS.items()
        }
        existing = session.query(ModelResult).filter_by(calibration_period_id=period_id).first()
        if existing:
            for key, value in payload.items():
                setattr(existing, key, value)
            updated += 1
        else:
            session.add(ModelResult(calibration_period_id=period_id, **payload))
            created += 1
    return created, updated


def seed_poco_do_fumo(db_path: str | None = None):
    engine, Session = initialize_db(db_path)
    session = Session()

    created = {
        "bacia": 0,
        "periodos": 0,
        "resultados": 0,
        "evap": 0,
        "chuva": 0,
        "vazao": 0,
    }
    updated = {
        "bacia": 0,
        "periodos": 0,
        "resultados": 0,
        "evap": 0,
        "chuva": 0,
        "vazao": 0,
    }
    missing_params = []

    try:
        parametros, missing_params = extract_bacia_parameters(XLSM_PATH)
        bacia = upsert_bacia(session, parametros)
        session.flush()
        bacia_id = cast(int, bacia.id)

        evap = read_monthly_evaporation(EVAP_PATH)
        chuva = read_daily_series(PLU_PATH)
        vazao = read_daily_series(FLU_PATH)

        c, u = upsert_monthly(session, bacia_id, evap)
        created["evap"] += c
        updated["evap"] += u

        c, u = upsert_daily(session, PrecipitationDaily, bacia_id, chuva)
        created["chuva"] += c
        updated["chuva"] += u

        c, u = upsert_daily(session, FlowDaily, bacia_id, vazao)
        created["vazao"] += c
        updated["vazao"] += u

        periodos_df = load_periods(PERIODOS_CSV)
        c, u = upsert_periods(session, bacia_id, periodos_df)
        created["periodos"] += c
        updated["periodos"] += u

        resultados_df = load_results(RESULTADOS_CSV)
        c, u = upsert_results(session, bacia_id, resultados_df)
        created["resultados"] += c
        updated["resultados"] += u

        session.commit()
    except Exception:
        session.rollback()
        raise
    finally:
        session.close()
        engine.dispose()

    return {
        "created": created,
        "updated": updated,
        "missing_params": missing_params,
    }


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Seed específico da bacia Poço do Fumo")
    parser.add_argument("--db", default=None, help="Caminho opcional para o banco SQLite")
    args = parser.parse_args()

    result = seed_poco_do_fumo(args.db)
    print("Poço do Fumo seed finalizado")
    print("Criados:", result["created"])
    print("Atualizados:", result["updated"])
    if result["missing_params"]:
        print("Parâmetros não encontrados no XLSM:", ", ".join(result["missing_params"]))
